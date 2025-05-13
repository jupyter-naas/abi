from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
import json
import sys

# Add the project root to the Python path
project_root = Path(__file__).parent.parent.parent.parent.parent
sys.path.append(str(project_root))

def generate_usage_sheet():
    """Generate an Excel sheet with GPT-4o usage data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "GPT-4o Usage Data"
    
    # Set up headers
    headers = ['Date', 'Conversation ID', 'Topic', 'Message Count', 'Tokens Used', 'Processing Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Load and analyze message data
    messages_dir = project_root / 'storage' / 'datastore' / 'gpt-4o' / 'messages'
    message_files = list(messages_dir.glob('*.json'))
    
    # Add data rows
    row = 2
    for msg_file in sorted(message_files, key=lambda x: x.stat().st_mtime):
        with open(msg_file, 'r') as f:
            msg_data = json.load(f)
            
            # Extract data
            date = datetime.fromtimestamp(msg_file.stat().st_mtime).strftime('%Y-%m-%d')
            conv_id = msg_file.stem
            topic = msg_data.get('topic', 'N/A')
            msg_count = len(msg_data.get('messages', []))
            tokens = msg_data.get('total_tokens', 0)
            proc_time = msg_data.get('processing_time', 0)
            
            # Add row data
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=conv_id)
            ws.cell(row=row, column=3, value=topic)
            ws.cell(row=row, column=4, value=msg_count)
            ws.cell(row=row, column=5, value=tokens)
            ws.cell(row=row, column=6, value=proc_time)
            
            row += 1
    
    # Add summary statistics
    ws.cell(row=row+1, column=1, value="Summary Statistics").font = Font(bold=True)
    ws.cell(row=row+2, column=1, value="Total Conversations")
    ws.cell(row=row+2, column=2, value=len(message_files))
    ws.cell(row=row+3, column=1, value="Average Messages per Conversation")
    ws.cell(row=row+3, column=2, value=f"=AVERAGE(D2:D{row})")
    ws.cell(row=row+4, column=1, value="Total Tokens Used")
    ws.cell(row=row+4, column=2, value=f"=SUM(E2:E{row})")
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    output_dir = project_root / 'storage' / 'datastore' / 'analytics' / 'sheets'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f'gpt4o_usage_data_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(output_file)
    
    return output_file

if __name__ == "__main__":
    sheet_file = generate_usage_sheet()
    print(f"Generated usage sheet: {sheet_file}")
