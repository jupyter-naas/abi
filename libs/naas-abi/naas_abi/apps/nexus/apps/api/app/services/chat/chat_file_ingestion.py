from __future__ import annotations

import hashlib
import io
import re
import uuid
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import PurePosixPath
from xml.etree import ElementTree as ET

import numpy as np
from naas_abi.apps.nexus.apps.api.app.services.chat.chat_file_embeddings import (
    build_chat_collection_name,
    build_embedding_cache_key,
    chunk_markdown,
    embed_texts,
)
from naas_abi_core.services.cache.CacheService import CacheService
from naas_abi_core.services.object_storage.ObjectStoragePort import Exceptions
from naas_abi_core.services.object_storage.ObjectStorageService import ObjectStorageService
from naas_abi_core.services.vector_store.VectorStoreService import VectorStoreService

# Callback type: (stage_name, overall_progress_pct 0-100) -> None
ProgressCallback = Callable[[str, int], None]

# PDF conversion: progress spans 30 % → 70 % of the overall job
_PDF_PROGRESS_START = 30
_PDF_PROGRESS_END = 70
# Number of progress updates to emit during PDF conversion (approx)
_PDF_PROGRESS_STEPS = 20


class ChatFileIngestionError(Exception):
    pass


@dataclass(frozen=True)
class ChatFileIngestionResult:
    conversation_id: str
    source_path: str
    collection_name: str
    file_sha256: str
    cache_hit: bool
    chunks_count: int
    statuses: list[str]
    embedding_model: str
    embedding_dimension: int


class ChatFileIngestionService:
    def __init__(
        self,
        object_storage: ObjectStorageService,
        vector_store: VectorStoreService,
        cache_service: CacheService,
    ):
        self.object_storage = object_storage
        self.vector_store = vector_store
        self.cache_service = cache_service

    @staticmethod
    def my_drive_uploads_path(user_id: str) -> str:
        return PurePosixPath("my-drive", user_id, "uploads").as_posix()

    def upload_and_ingest(
        self,
        user_id: str,
        conversation_id: str,
        filename: str,
        content: bytes,
        embedding_model: str = "hash-v1",
        embedding_dimension: int = 256,
        on_progress: ProgressCallback | None = None,
    ) -> ChatFileIngestionResult:
        safe_filename = PurePosixPath(filename or "untitled").name
        if not safe_filename:
            raise ChatFileIngestionError("Invalid file name")

        upload_prefix = self.my_drive_uploads_path(user_id)
        self.object_storage.put_object(upload_prefix, safe_filename, content)
        source_path = PurePosixPath(upload_prefix, safe_filename).as_posix()
        return self.ingest_from_path(
            user_id=user_id,
            conversation_id=conversation_id,
            source_path=source_path,
            embedding_model=embedding_model,
            embedding_dimension=embedding_dimension,
            on_progress=on_progress,
        )

    def ingest_from_path(
        self,
        user_id: str,
        conversation_id: str,
        source_path: str,
        embedding_model: str = "hash-v1",
        embedding_dimension: int = 256,
        on_progress: ProgressCallback | None = None,
    ) -> ChatFileIngestionResult:
        statuses = ["hashing", "cache_lookup"]

        # Normalise the path to resolve any ".." traversal attempts before
        # checking ownership (e.g. "my-drive/user/../../etc/passwd" must not pass).
        import posixpath
        normalized_source = posixpath.normpath(source_path.strip().strip("/"))
        expected_prefix = PurePosixPath("my-drive", user_id).as_posix()
        if not (
            normalized_source == expected_prefix
            or normalized_source.startswith(f"{expected_prefix}/")
        ):
            raise ChatFileIngestionError("Source path must belong to user's My Drive")

        prefix, key = self._split_source(normalized_source)
        try:
            file_content = self.object_storage.get_object(prefix, key)
        except Exceptions.ObjectNotFound as exc:
            raise ChatFileIngestionError("File not found") from exc

        file_sha = hashlib.sha256(file_content).hexdigest()
        cache_key = build_embedding_cache_key(file_sha, embedding_model, embedding_dimension)
        collection_name = build_chat_collection_name(conversation_id)

        cache_hit = False
        markdown = ""
        chunks: list[str] = []
        vectors: list[np.ndarray] = []

        if self.cache_service.exists(cache_key):
            cached = self.cache_service.get(cache_key)
            cached_chunks = cached.get("chunks") if isinstance(cached, dict) else None
            cached_vectors = cached.get("vectors") if isinstance(cached, dict) else None
            if isinstance(cached_chunks, list) and isinstance(cached_vectors, list):
                chunks = [str(chunk) for chunk in cached_chunks]
                vectors = [np.array(v, dtype=float) for v in cached_vectors]
                cache_hit = True
                statuses.append("reusing_vectors")
                if on_progress:
                    on_progress("reusing_vectors", 90)

        if not cache_hit:
            statuses.extend(["converting", "chunking", "vectorizing"])

            # --- Converting ---
            if on_progress:
                on_progress("converting", _PDF_PROGRESS_START)
            markdown = self._to_markdown(normalized_source, file_content, on_progress=on_progress)
            if not markdown.strip():
                raise ChatFileIngestionError("Empty content extracted from file")

            # --- Chunking ---
            if on_progress:
                on_progress("chunking", 72)
            chunks = chunk_markdown(markdown)

            # --- Vectorizing ---
            if on_progress:
                on_progress("vectorizing", 75)

            def _embed_progress(batch_pct: int) -> None:
                # Map 0-100 embed progress → 75-90% overall
                if on_progress:
                    on_progress("vectorizing", 75 + int(batch_pct * 0.15))

            vectors = embed_texts(
                chunks,
                embedding_model,
                embedding_dimension,
                on_progress=_embed_progress if on_progress else None,
            )

            # --- Cache write ---
            payload = {
                "schema_version": 1,
                "file_sha256": file_sha,
                "embedding_model": embedding_model,
                "embedding_dimension": embedding_dimension,
                "chunking_strategy_version": "char-v1",
                "markdown_content": markdown,
                "chunks": chunks,
                "vectors": [v.tolist() for v in vectors],
                "created_at": datetime.now(UTC).isoformat(),
            }
            self.cache_service.set_json_if_absent(cache_key, payload)

        # "upserting" progress applies to both cache-hit and cache-miss paths.
        # Emit it here so the UI always shows the upsert stage while batches
        # are written to the vector store.
        if on_progress:
            on_progress("upserting", 92)

        self._upsert_chunks(
            collection_name=collection_name,
            chunks=chunks,
            vectors=vectors,
            conversation_id=conversation_id,
            user_id=user_id,
            source_path=normalized_source,
            file_sha256=file_sha,
            embedding_model=embedding_model,
            embedding_dimension=embedding_dimension,
            cache_hit=cache_hit,
            on_progress=on_progress,
        )

        statuses.append("ready")
        return ChatFileIngestionResult(
            conversation_id=conversation_id,
            source_path=normalized_source,
            collection_name=collection_name,
            file_sha256=file_sha,
            cache_hit=cache_hit,
            chunks_count=len(chunks),
            statuses=statuses,
            embedding_model=embedding_model,
            embedding_dimension=embedding_dimension,
        )

    @staticmethod
    def _split_source(source_path: str) -> tuple[str, str]:
        if "/" not in source_path:
            return "", source_path
        prefix, key = source_path.rsplit("/", 1)
        return prefix, key

    def _upsert_chunks(
        self,
        collection_name: str,
        chunks: list[str],
        vectors: list[np.ndarray],
        conversation_id: str,
        user_id: str,
        source_path: str,
        file_sha256: str,
        embedding_model: str,
        embedding_dimension: int,
        cache_hit: bool,
        on_progress: ProgressCallback | None = None,
    ) -> None:
        if not chunks:
            raise ChatFileIngestionError("No chunks generated from content")
        if len(chunks) != len(vectors):
            raise ChatFileIngestionError("Chunk/vector mismatch")

        self.vector_store.ensure_collection(
            collection_name=collection_name,
            dimension=embedding_dimension,
            distance_metric="cosine",
        )

        # Qdrant has a default payload size limit of 32 MB per upsert request.
        # For large documents (hundreds of pages) a single batch of all chunks
        # easily exceeds that: 1536-dim vectors alone are ~12 bytes/float as
        # JSON, so 2000 chunks × 1536 dims ≈ 36 MB.  We therefore upsert in
        # fixed-size batches that keep each request well under the limit.
        _UPSERT_BATCH = 100

        now_iso = datetime.now(UTC).isoformat()
        filename = PurePosixPath(source_path).name

        # Pre-build all IDs, metadata, and payloads.
        ids: list[str] = []
        metadata: list[dict] = []
        payloads: list[dict] = []

        for index, chunk in enumerate(chunks):
            # Qdrant requires point IDs to be unsigned integers or UUIDs.
            # Derive a deterministic UUID from the SHA-256 of the chunk key.
            raw = hashlib.sha256(
                f"{conversation_id}:{file_sha256}:{embedding_model}:{embedding_dimension}:{index}".encode()
            ).hexdigest()
            doc_id = str(uuid.UUID(hex=raw[:32]))
            ids.append(doc_id)
            # NOTE: do NOT store chunk_text in metadata — it would duplicate the
            # text that is already stored in the payload ({"text": chunk}) and
            # double the request size.  The search service reads payload["text"]
            # first and only falls back to metadata["chunk_text"] for legacy points.
            metadata.append(
                {
                    "thread_id": conversation_id,
                    "user_id": user_id,
                    "source_path": source_path,
                    "filename": filename,
                    "file_sha256": file_sha256,
                    "embedding_model": embedding_model,
                    "embedding_dimension": embedding_dimension,
                    "cache_hit": cache_hit,
                    "chunk_index": index,
                    "ingested_at": now_iso,
                }
            )
            payloads.append({"text": chunk})

        # Upsert in batches to stay under Qdrant's payload size limit.
        # Report progress from 92 % → 99 % so the UI shows movement while
        # batches are written (the worker emits the final 100 % as "ready").
        total = len(ids)
        for start in range(0, total, _UPSERT_BATCH):
            end = min(start + _UPSERT_BATCH, total)
            self.vector_store.add_documents(
                collection_name=collection_name,
                ids=ids[start:end],
                vectors=vectors[start:end],
                metadata=metadata[start:end],
                payloads=payloads[start:end],
            )
            if on_progress:
                # Map batch completion (0-100) to 92-99 % overall.
                batch_pct = int(end / total * 100)
                on_progress("upserting", 92 + int(batch_pct * 0.07))

    def _to_markdown(
        self,
        source_path: str,
        content: bytes,
        on_progress: ProgressCallback | None = None,
    ) -> str:
        ext = PurePosixPath(source_path).suffix.lower()
        if ext in {".md", ".txt", ".py", ".json", ".yaml", ".yml", ".csv"}:
            return content.decode("utf-8")
        if ext == ".docx":
            return self._docx_to_markdown(content)
        if ext == ".pptx":
            return self._pptx_to_markdown(content)
        if ext == ".xlsx":
            return self._xlsx_to_markdown(content)
        if ext == ".pdf":
            return self._pdf_to_markdown(content, on_progress=on_progress)
        raise ChatFileIngestionError(f"Unsupported file type: {ext}")

    @staticmethod
    def _pdf_to_markdown(
        content: bytes,
        on_progress: ProgressCallback | None = None,
    ) -> str:
        """Convert a PDF to Markdown.

        When *on_progress* is supplied and the document has more than
        ``_PDF_PROGRESS_STEPS`` pages, the conversion is performed in batches
        so the caller receives incremental progress events in the
        ``_PDF_PROGRESS_START`` → ``_PDF_PROGRESS_END`` percent range.
        """
        try:
            import pymupdf4llm
        except ImportError as exc:
            raise ChatFileIngestionError("PDF support requires pymupdf4llm") from exc

        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".pdf") as tmp:
            tmp.write(content)
            tmp.flush()

            if on_progress is not None:
                # Try to count pages and process in batches for granular progress.
                try:
                    import fitz  # PyMuPDF — always available as a pymupdf4llm dep

                    with fitz.open(tmp.name) as doc:
                        total_pages = len(doc)

                    if total_pages > _PDF_PROGRESS_STEPS:
                        batch_size = max(1, total_pages // _PDF_PROGRESS_STEPS)
                        parts: list[str] = []
                        for start in range(0, total_pages, batch_size):
                            end = min(start + batch_size, total_pages)
                            pages = list(range(start, end))
                            part = str(pymupdf4llm.to_markdown(tmp.name, pages=pages))
                            parts.append(part)
                            pct = _PDF_PROGRESS_START + int(
                                (end / total_pages)
                                * (_PDF_PROGRESS_END - _PDF_PROGRESS_START)
                            )
                            on_progress("converting", pct)
                        return "\n\n".join(parts)
                except Exception:
                    # fitz not available or page-batch failed — fall through to
                    # the single-call path; progress will stall at 30 % until done.
                    pass

            return str(pymupdf4llm.to_markdown(tmp.name))

    @staticmethod
    def _docx_to_markdown(content: bytes) -> str:
        namespace = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        ns_map = {"w": namespace}

        try:
            with zipfile.ZipFile(io.BytesIO(content)) as docx_file:
                xml_content = docx_file.read("word/document.xml")
        except (KeyError, zipfile.BadZipFile) as exc:
            raise ChatFileIngestionError("Invalid DOCX content: missing word/document.xml") from exc

        root = ET.fromstring(xml_content)  # nosec B314 — XML from a trusted DOCX zip, no external entity expansion
        lines: list[str] = []
        for paragraph in root.findall(".//w:body/w:p", ns_map):
            style_elem = paragraph.find("./w:pPr/w:pStyle", ns_map)
            style = style_elem.get(f"{{{namespace}}}val", "") if style_elem is not None else ""

            text_parts: list[str] = []
            for node in paragraph.iter():
                if node.tag == f"{{{namespace}}}t" and node.text:
                    text_parts.append(node.text)
                elif node.tag == f"{{{namespace}}}tab":
                    text_parts.append("    ")
                elif node.tag in {f"{{{namespace}}}br", f"{{{namespace}}}cr"}:
                    text_parts.append("\n")

            text = re.sub(r"[ \t]+", " ", "".join(text_parts)).strip()
            if not text:
                continue

            heading_match = re.match(r"Heading([1-6])$", style)
            if heading_match:
                level = int(heading_match.group(1))
                lines.append(f"{'#' * level} {text}")
            elif paragraph.find("./w:pPr/w:numPr", ns_map) is not None:
                lines.append(f"- {text}")
            else:
                lines.append(text)

        return "\n\n".join(lines).strip()

    @staticmethod
    def _pptx_to_markdown(content: bytes) -> str:
        p_ns = "http://schemas.openxmlformats.org/presentationml/2006/main"
        a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        ns_map = {"p": p_ns, "a": a_ns}

        try:
            with zipfile.ZipFile(io.BytesIO(content)) as pptx_file:
                slide_paths = [
                    name
                    for name in pptx_file.namelist()
                    if name.startswith("ppt/slides/slide") and name.endswith(".xml")
                ]
                slide_paths.sort(
                    key=lambda path: int(re.search(r"slide(\d+)\.xml$", path).group(1))
                    if re.search(r"slide(\d+)\.xml$", path)
                    else 0
                )

                markdown_parts: list[str] = []
                for slide_index, slide_path in enumerate(slide_paths, start=1):
                    slide_xml = pptx_file.read(slide_path)
                    root = ET.fromstring(slide_xml)  # nosec B314 — XML from a trusted PPTX zip, no external entity expansion
                    lines: list[str] = []
                    for paragraph in root.findall(".//a:p", ns_map):
                        parts: list[str] = []
                        for node in paragraph.iter():
                            if node.tag == f"{{{a_ns}}}t" and node.text:
                                parts.append(node.text)
                            elif node.tag == f"{{{a_ns}}}br":
                                parts.append("\n")
                        text = re.sub(r"[ \t]+", " ", "".join(parts)).strip()
                        if text:
                            lines.append(text)

                    if lines:
                        markdown_parts.append(f"## Slide {slide_index}")
                        markdown_parts.extend(f"- {line}" for line in lines)
                return "\n\n".join(markdown_parts).strip()
        except (KeyError, zipfile.BadZipFile) as exc:
            raise ChatFileIngestionError("Invalid PPTX content: unable to read slide XML") from exc

    @staticmethod
    def _xlsx_to_markdown(content: bytes) -> str:
        try:
            import openpyxl
        except ImportError as exc:
            raise ChatFileIngestionError("XLSX support requires openpyxl") from exc

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ChatFileIngestionError("Invalid XLSX content") from exc

        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Drop trailing all-None rows
            while rows and all(cell is None for cell in rows[-1]):
                rows.pop()
            if not rows:
                continue

            parts.append(f"## {sheet_name}")

            # Determine column count from the widest row
            col_count = max(len(row) for row in rows)

            def cell_str(val: object) -> str:
                if val is None:
                    return ""
                return re.sub(r"[\|\n\r]", " ", str(val)).strip()

            header, *data_rows = rows
            header_cells = [cell_str(v) for v in header] + [""] * (col_count - len(header))
            separator = ["---"] * col_count

            table_lines = [
                "| " + " | ".join(header_cells) + " |",
                "| " + " | ".join(separator) + " |",
            ]
            for row in data_rows:
                cells = [cell_str(v) for v in row] + [""] * (col_count - len(row))
                table_lines.append("| " + " | ".join(cells) + " |")

            parts.append("\n".join(table_lines))

        wb.close()
        if not parts:
            raise ChatFileIngestionError("Empty content extracted from XLSX")
        return "\n\n".join(parts).strip()
